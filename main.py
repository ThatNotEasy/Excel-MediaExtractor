import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from zipfile import ZipFile
import shutil
from PIL import Image as PILImage
import io

def extract_tables(excel_file, output_folder):
    df_dict = pd.read_excel(excel_file, sheet_name=None)
    for sheet, df in df_dict.items():
        csv_path = os.path.join(output_folder, f"{sheet}.csv")
        df.to_csv(csv_path, index=False)
        print(f"Extracted table saved: {csv_path}")

def extract_images(excel_file, output_folder):
    wb = load_workbook(excel_file)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for img_index, image in enumerate(sheet._images):
            if isinstance(image, Image):
                img_data = image._data()
                img = PILImage.open(io.BytesIO(img_data))
                img_path = os.path.join(output_folder, f"{sheet_name}_image_{img_index + 1}.png")
                img.save(img_path)
                print(f"Extracted image saved: {img_path}")

def extract_embedded_files(excel_file, output_folder):
    temp_folder = os.path.join(output_folder, "temp")
    os.makedirs(temp_folder, exist_ok=True)
    
    with ZipFile(excel_file, 'r') as zip_ref:
        zip_ref.extractall(temp_folder)
    
    ole_objects_folder = os.path.join(temp_folder, "xl", "embeddings")
    if os.path.exists(ole_objects_folder):
        for file in os.listdir(ole_objects_folder):
            if file.endswith(".bin"):
                file_path = os.path.join(ole_objects_folder, file)
                new_file_path = os.path.join(output_folder, file.replace(".bin", ".pdf"))
                shutil.move(file_path, new_file_path)
                print(f"Extracted embedded file saved: {new_file_path}")
    
    shutil.rmtree(temp_folder)  # Clean up temp files

def extract_from_excel(excel_file, output_folder):
    os.makedirs(output_folder, exist_ok=True)
    extract_tables(excel_file, output_folder)
    extract_images(excel_file, output_folder)
    extract_embedded_files(excel_file, output_folder)

# Example usage
excel_file_path = "example.xlsx"
output_directory = "output"
extract_from_excel(excel_file_path, output_directory)
